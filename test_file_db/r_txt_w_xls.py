from openpyxl import Workbook
import openpyxl
import os, sys

# Проверяем наличие файла
os.chdir(r"D:\python_lab\test_file_db")
if os.path.exists('text_01.txt') == True:
    print("Файл *.txt существует")
else:
    print("Файл *.txt НЕ существует, операцию не возможно выполнить!")

# s = {'ФИО': [номер счета, сумма на счете]} - создаем пустой словарь
s = {}

# начинаем открывать файл с данными и записываем в словарь
f = open('text_01.txt', 'r', encoding='utf-8')
line = f.readline()

while line:
    # print("\nНомер счета: " + line[0:11])
    # print("\tФИО владельца счета: " + line[11:41])
    # print("\tСумма на счете: " + line[41:47])
    a = line[0:11]
    b = line[11:41]
    c = line[41:47]
    s[b.strip()] = [a, c]  # тут убираем все лишние пробелы вокруг ФИО
    line = f.readline()
    f.close

# Создаем Excel файл
wb = Workbook()
ws = wb.create_sheet('Data', 0)
wb.active = 0
# print(wb.sheetnames) # Просмотр всех листов (название)
# print(wb.active) # Просмотр активного листа
row = 1
wb.active["A"+str(row)] = 'FIO'
wb.active["B"+str(row)] = 'Number bank'
wb.active["C"+str(row)] = 'Money'

# Записываем в Excel данные из словаря
for x, y in s.items():
    row += 1
    wb.active["A"+str(row)] = x
    wb.active["B"+str(row)] = y[0]
    wb.active["C"+str(row)] = y[1]

wb.save('filename.xlsx')
